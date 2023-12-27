# Importa las bibliotecas necesarias
from django.http import HttpResponse
from django.shortcuts import redirect, get_object_or_404
from django.template import loader
from openpyxl import Workbook
from rest_framework import viewsets, permissions

from jugadores.forms import JugadorFormulario
from jugadores.models import Jugador, Plataforma, Modo, GeneroJuego, Juego
from jugadores.serializers import JugadorSerializer, PlataformaSerializer, ModoSerializer, GeneroJuegoSerializer, \
    JuegoSerializer


# Vista para agregar un jugador
def agregar_jugador(request):
    pagina = loader.get_template('agregar_jugador.html')
    if request.method == 'GET':
        formulario = JugadorFormulario()  # Corregir esta línea
    elif request.method == 'POST':
        formulario = JugadorFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('jugadores')
    datos = {'formulario':formulario}
    return HttpResponse(pagina.render(datos, request))


# Vista para modificar un jugador
def modificar_jugador(request, jugador_id):
    pagina = loader.get_template('modificar_jugador.html')
    jugador = get_object_or_404(Jugador, pk=jugador_id)
    if request.method == 'GET':
        formulario = JugadorFormulario(instance=jugador)
    elif request.method == 'POST':
        formulario = JugadorFormulario(request.POST, instance=jugador)
        if formulario.is_valid():
            formulario.save()
            return redirect('jugadores')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))


# Vista para ver un jugador
def ver_jugador(request, jugador_id):
    jugador = get_object_or_404(Jugador, pk=jugador_id)
    datos = {'jugador': jugador}
    pagina = loader.get_template('ver_jugador.html')
    return HttpResponse(pagina.render(datos, request))


# Vista para eliminar un jugador
def eliminar_jugador(request, jugador_id):
    jugador = get_object_or_404(Jugador, pk=jugador_id)
    if jugador:
        jugador.delete()
        return redirect('jugadores')


# Vista para generar un informe de jugadores en formato Excel
def generar_reporte_jugadores(request, *args, **kwargs):
    jugadores = Jugador.objects.order_by('nombre', 'sexo', 'generos_preferidos', 'plataforma_preferida',
                                         'modos_preferidos')
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'REPORTE DE JUGADORES'
    ws.merge_cells('B1:G1')
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'SEXO'
    ws['D3'] = 'GENERO PREFERIDO'
    ws['E3'] = 'PLATAFORMA PREFERIDA'
    ws['F3'] = 'MODO PREFERIDO'
    cont = 5

    for jugador in jugadores:
        ws.cell(row=cont, column=2).value = jugador.nombre
        ws.cell(row=cont, column=3).value = jugador.sexo
        ws.cell(row=cont, column=4).value = str(jugador.generos_preferidos)
        ws.cell(row=cont, column=5).value = str(jugador.plataforma_preferida)
        ws.cell(row=cont, column=6).value = str(jugador.modos_preferidos)
        cont += 1

    nombre_archivo = "ReporteJugadoresExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


class JugadorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Jugador.objects.all().order_by('-nombre')
    serializer_class = JugadorSerializer
    permission_classes = [permissions.IsAuthenticated]

class  JuegoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Juego.objects.all()
    serializer_class =  JuegoSerializer
    permission_classes = [permissions.IsAuthenticated]
class PlataformaViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Plataforma.objects.all()
    serializer_class = PlataformaSerializer
    permission_classes = [permissions.IsAuthenticated]

class ModoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Modo.objects.all()
    serializer_class = ModoSerializer
    permission_classes = [permissions.IsAuthenticated]

class GeneroJuegoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = GeneroJuego.objects.all()
    serializer_class = GeneroJuegoSerializer
    permission_classes = [permissions.IsAuthenticated]
